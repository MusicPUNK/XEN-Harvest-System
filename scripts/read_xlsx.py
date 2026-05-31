#!/usr/bin/env python3
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

ADDRESS_RE = re.compile(r"0x[a-fA-F0-9]{40}")


def cell_value(row, index):
    return row[index] if index < len(row) else None


def main():
    if len(sys.argv) != 2:
        raise SystemExit("Usage: read_xlsx.py <workbook.xlsx>")

    path = Path(sys.argv[1])
    wb = load_workbook(path, read_only=True, data_only=True)
    output = []

    for ws in wb.worksheets:
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        header_text = " ".join("" if value is None else str(value) for value in header)
        wallet_match = ADDRESS_RE.search(header_text)
        wallet = wallet_match.group(0) if wallet_match else ""

        for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            label = cell_value(row, 0)
            quantity = cell_value(row, 4)
            if label is None or str(label).strip() == "":
                continue
            if quantity is None and cell_value(row, 1) is None and cell_value(row, 3) is None:
                continue
            output.append(
                {
                    "sheet": ws.title,
                    "rowNumber": row_number,
                    "wallet": wallet,
                    "label": label,
                    "mintDateRaw": cell_value(row, 1),
                    "termDaysRaw": cell_value(row, 2),
                    "expiryRaw": cell_value(row, 3),
                    "quantityRaw": quantity,
                    "claimAmountRaw": cell_value(row, 11),
                }
            )

    print(json.dumps(output, ensure_ascii=False, default=str))


if __name__ == "__main__":
    main()
